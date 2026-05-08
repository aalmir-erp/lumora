"""Desktop control: full-screen screenshots + mouse/keyboard via pyautogui.

Plus structured Excel I/O via openpyxl (preferred over GUI clicks for data tasks).
"""
from __future__ import annotations

import base64
import io
import os
from typing import Any

from .safety import confirm_action


def screenshot_b64() -> str:
    import mss  # type: ignore
    from PIL import Image  # type: ignore
    with mss.mss() as sct:
        m = sct.monitors[1]
        raw = sct.grab(m)
        img = Image.frombytes("RGB", raw.size, raw.rgb)
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        return base64.b64encode(buf.getvalue()).decode()


def click(x: int, y: int, button: str = "left") -> None:
    import pyautogui  # type: ignore
    if not confirm_action(f"click at ({x},{y}) {button}"):
        return
    pyautogui.click(x=x, y=y, button=button)


def type_text(text: str, delay_ms: int = 40) -> None:
    import pyautogui  # type: ignore
    if not confirm_action(f"type {text!r}"):
        return
    pyautogui.typewrite(text, interval=delay_ms / 1000.0)


def hotkey(keys: list[str]) -> None:
    import pyautogui  # type: ignore
    if not confirm_action(f"hotkey {'+'.join(keys)}"):
        return
    pyautogui.hotkey(*keys)


def excel_read(path: str, sheet: str | None = None, range_: str | None = None) -> list[list[Any]]:
    from openpyxl import load_workbook  # type: ignore
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    if range_:
        cells = ws[range_]
        return [[c.value for c in row] for row in cells]
    return [[c.value for c in row] for row in ws.iter_rows()]


def excel_write(path: str, sheet: str | None = None, cells: dict[str, Any] | None = None) -> dict[str, Any]:
    from openpyxl import load_workbook, Workbook  # type: ignore
    if os.path.exists(path):
        wb = load_workbook(path)
    else:
        wb = Workbook()
    if sheet and sheet in wb.sheetnames:
        ws = wb[sheet]
    elif sheet:
        ws = wb.create_sheet(sheet)
    else:
        ws = wb.active
    written = 0
    for ref, val in (cells or {}).items():
        ws[ref] = val
        written += 1
    wb.save(path)
    return {"written": written, "path": path, "sheet": ws.title}
